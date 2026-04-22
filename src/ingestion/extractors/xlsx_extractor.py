"""
xlsx_extractor.py — Extractor for .xlsx files.

Processes all sheets in a workbook. Converts each row into a natural
language sentence using the header row as column labels, prefixed with
the sheet name for context.
"""

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


def _cell_to_str(value) -> str:
    """
    Convert an Excel cell value to a clean string.

    Args:
        value: Raw cell value — could be str, int, float, date, bool, or None.

    Returns:
        String representation of the value, or empty string if None.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"       # True/False → Yes/No
    if isinstance(value, float):
        # drop .0 from whole numbers — 150.0 → "150"
        return str(int(value)) if value == int(value) else str(round(value, 4))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")      # consistent date format
    return str(value).strip()


def _row_to_sentence(sheet_name: str, headers: list[str], row) -> str:
    """
    Convert a single Excel row into a natural language sentence.

    Args:
        sheet_name: Name of the sheet this row belongs to.
        headers:    List of column header strings.
        row:        Tuple of openpyxl Cell objects for this row.

    Returns:
        A readable sentence prefixed with the sheet name.
    """
    pairs = []
    for header, cell in zip(headers, row):
        header = header.strip()
        value = _cell_to_str(cell.value)
        if header and value:                   # skip empty headers or cells
            pairs.append(f"{header} is {value}")

    if not pairs:
        return ""

    sentence = ". ".join(pairs) + "."
    return f"[{sheet_name}] {sentence}"        # prefix with sheet name


def extract(file_path: Path) -> str:
    """
    Extract text from all sheets in an Excel workbook.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        One sentence per data row across all sheets, joined by newlines.
        Returns empty string if no data rows are found.

    Raises:
        ValueError: If the file extension is not .xlsx.
    """
    if file_path.suffix.lower() != ".xlsx":
        raise ValueError(f"xlsx_extractor only handles .xlsx, got: {file_path.suffix}")

    # data_only=True returns computed cell values, not formula strings
    workbook = load_workbook(file_path, data_only=True)

    all_lines = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows())         # load all rows at once

        if len(rows) < 2:                      # need at least header + one data row
            continue

        # first row is headers — extract as strings
        headers = [_cell_to_str(cell.value) for cell in rows[0]]

        if not any(headers):                   # skip sheets with no header row
            continue

        sheet_lines = []
        for row in rows[1:]:                   # everything after the header
            # skip rows where all cells are empty
            if not any(cell.value is not None for cell in row):
                continue

            sentence = _row_to_sentence(sheet_name, headers, row)
            if sentence:
                sheet_lines.append(sentence)

        if sheet_lines:
            all_lines.append(f"=== Sheet: {sheet_name} ===")
            all_lines.extend(sheet_lines)

    workbook.close()

    if not all_lines:
        print(f"  [xlsx_extractor] Warning: no data rows found in {file_path.name}")
        return ""

    return "\n".join(all_lines)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python -m src.ingestion.extractors.xlsx_extractor <file.xlsx>")
        sys.exit(1)

    target = Path(sys.argv[1])

    if not target.exists():
        print(f"Error: file not found — {target}")
        sys.exit(1)

    result = extract(target)

    print(f"--- Extracted {len(result):,} characters from {target.name} ---\n")
    print(result[:500])